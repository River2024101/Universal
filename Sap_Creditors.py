# ============================================================
# SAP OS AGEING PROCESSOR - NEW EXCEL FORMAT
# ============================================================
# Purpose:
# 1. Read new SapOS.xlsx format where the same header repeats supplier-wise.
# 2. Ignore blank rows, repeated headers, and subtotal rows marked with * / **.
# 3. Create clean final SAP_OS output with ageing / cash-flow bucket.
# 4. Read Supplier Master for Group and View - Y/N.
# 5. Add missing suppliers to master in red font.
# ============================================================

import os
import re
from datetime import datetime, date
from decimal import Decimal, getcontext
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# ================== PATHS ==================
# Change only these 3 paths as per your computer folder.
INPUT_PATH = r"D:\New Sap Cr Ageing\SapOS.xlsx"
MASTER_PATH = r"D:\New Sap Cr Ageing\Sap_CreditorMaster.xlsx"
OUTPUT_PATH = r"D:\New Sap Cr Ageing\SapOS_Final_Output.xlsx"

# Sheet name in the new SAP OS file. Keep blank/None to auto-pick first sheet.
INPUT_SHEET_NAME = "Data"

getcontext().prec = 60

# ================== COMMON HELPERS ==================

def normalize_text(x):
    """Return clean text; blank/NaN/None becomes empty string."""
    if x is None:
        return ""
    s = str(x).strip()
    if s.lower() in ("", "nan", "none", "nat"):
        return ""
    return s


def normalize_supplier_code(x):
    """Convert supplier code like 2300001.0 to 2300001 and keep text safely."""
    s = normalize_text(x)
    if not s:
        return ""
    try:
        if re.fullmatch(r"\d+(\.0+)?", s):
            return str(int(float(s)))
    except Exception:
        pass
    return s


def is_valid_supplier_code(x):
    s = normalize_supplier_code(x)
    return bool(s) and s.isdigit() and len(s) >= 4


SCI_RE = re.compile(r"^[+-]?\d+(\.\d+)?[eE][+-]?\d+$")


def normalize_reference(x):
    """Avoid scientific notation for long reference/document numbers."""
    s = normalize_text(x)
    if not s:
        return ""
    if SCI_RE.match(s):
        try:
            d = Decimal(s)
            return format(d.to_integral_value(), "f")
        except Exception:
            return s
    try:
        if re.fullmatch(r"\d+\.0+", s):
            return str(int(float(s)))
    except Exception:
        pass
    return s


def clean_amount(x):
    """Convert SAP amount strings to number. Handles commas, spaces, and trailing minus."""
    s = normalize_text(x)
    if not s:
        return None

    # SAP sometimes gives 1,234.00- for negative amount
    trailing_minus = s.endswith("-")
    if trailing_minus:
        s = s[:-1]

    # Accounting format: (1,234.00)
    bracket_negative = s.startswith("(") and s.endswith(")")
    if bracket_negative:
        s = s[1:-1]

    s = s.replace(",", "").replace(" ", "")
    num = pd.to_numeric(s, errors="coerce")
    if pd.isna(num):
        return None

    num = float(num)
    if trailing_minus or bracket_negative:
        num = -abs(num)
    return num


def parse_date_safe(x):
    """Parse dd.mm.yyyy / dd-mm-yyyy / Excel date safely."""
    if x is None or str(x).strip().lower() in ("", "nan", "none", "nat"):
        return None

    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x

    dt = pd.to_datetime(x, dayfirst=True, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()


# ================== DUE DATE / CASH FLOW LOGIC ==================

DD_TOKEN_RE = re.compile(r"\(.*?DD\s*:\s*([^)]+)\)", re.IGNORECASE)


def due_date_from_dd_value(dd_value):
    """
    Convert DD token into due date.
    Expected SAP text examples:
    DD:240526 => 24.05.2026
    DD:STD    => standard due date, use Pmnt Date instead
    """
    token = normalize_text(dd_value)
    if not token:
        return None

    token_no_space = re.sub(r"\s+", "", token).upper()
    token_letters_only = re.sub(r"[^A-Z]", "", token_no_space)

    if token_letters_only == "STD":
        return None

    digits = re.sub(r"\D", "", token_no_space)
    if len(digits) != 6:
        return None

    try:
        dd = int(digits[0:2])
        mm = int(digits[2:4])
        yy = int(digits[4:6])
        return date(2000 + yy, mm, dd)
    except Exception:
        return None


def parse_real_due_dt(text_value, pmnt_date_only, dd_column_value=None):
    """
    Priority:
    1. Due date from DD column if available and valid.
    2. Due date from Text value like (DD:240526).
    3. Pmnt Date.
    """
    due_from_dd_col = due_date_from_dd_value(dd_column_value)
    if due_from_dd_col:
        return due_from_dd_col

    txt = normalize_text(text_value)
    if txt:
        m = DD_TOKEN_RE.search(txt)
        if m:
            due_from_text = due_date_from_dd_value(m.group(1))
            if due_from_text:
                return due_from_text

    return pmnt_date_only


def cashflow_bucket(due_dt):
    """Bucket based on RealDue_Dt against today's date."""
    if due_dt is None:
        return "NO DUE DATE"

    today = date.today()
    delta = (due_dt - today).days

    if delta < 0:
        return "OVERDUE"
    if delta <= 7:
        return "DUE 0-7 DAYS"
    if delta <= 15:
        return "DUE 8-15 DAYS"
    if delta <= 30:
        return "DUE 16-30 DAYS"
    if delta <= 60:
        return "DUE 31-60 DAYS"
    return "DUE 61+ DAYS"


# ================== MASTER FILE ==================

def read_master_maps(master_path):
    """
    Read master file and return:
    - workbook / worksheet
    - column map
    - supplier-wise Group
    - supplier-wise View - Y/N
    - existing supplier set
    """
    if not Path(master_path).exists():
        raise FileNotFoundError(f"Master file not found: {master_path}")

    wb = load_workbook(master_path)
    ws = wb.active

    header_row = None
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, min(ws.max_column, 80) + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip().lower() == "supplier":
                header_row = r
                break
        if header_row:
            break

    if not header_row:
        raise Exception("Supplier column not found in master file. Please check master header.")

    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            col_map[str(v).strip().lower()] = c

    sup_c = col_map.get("supplier")
    name_c = col_map.get("name") or col_map.get("vendor name")
    group_c = col_map.get("group")
    view_c = col_map.get("view - y/n") or col_map.get("view-y/n") or col_map.get("view")

    if not sup_c:
        raise Exception("Supplier column missing in master file.")

    group_map = {}
    view_map = {}
    name_map = {}
    existing = set()

    for r in range(header_row + 1, ws.max_row + 1):
        sup = normalize_supplier_code(ws.cell(r, sup_c).value)
        if sup:
            existing.add(sup)
            group_map[sup] = ws.cell(r, group_c).value if group_c else ""
            view_map[sup] = ws.cell(r, view_c).value if view_c else ""
            name_map[sup] = ws.cell(r, name_c).value if name_c else ""

    return wb, ws, col_map, group_map, view_map, name_map, existing


def append_missing_to_master(ws, col_map, missing_rows):
    """Append missing suppliers in red font."""
    if not missing_rows:
        return

    red = Font(color="FF0000")

    sup_c = col_map.get("supplier", 1)
    name_c = col_map.get("name") or col_map.get("vendor name") or 2
    added_c = col_map.get("added on") or col_map.get("added date") or None

    r = ws.max_row + 1
    for sup, name, added_on in missing_rows:
        ws.cell(r, sup_c, sup).font = red
        ws.cell(r, name_c, name).font = red
        if added_c:
            ws.cell(r, added_c, added_on).font = red
        r += 1


# ================== NEW SAP OS EXCEL READER ==================

EXPECTED_HEADERS = [
    "Supplier", "Vendor Name", "DocumentNo", "St", "Reference", "PK", "Type", "DD",
    "Doc. Date", "Pstng Date", "S", "Pmnt Date", "Local Crcy Amt", "LCu",
    "Amount in DC", "Crcy", "Text", "User Name", "Clearing"
]


def find_header_positions(raw_df):
    """
    Find repeated header rows and identify the columns based on the row where Supplier/Vendor Name exists.
    Returns list of (row_index, col_map_for_input).
    """
    positions = []

    for idx in range(len(raw_df)):
        row_values = [normalize_text(v) for v in raw_df.iloc[idx].tolist()]
        lowered = [v.lower() for v in row_values]

        if "supplier" in lowered and "vendor name" in lowered and "documentno" in lowered:
            input_col_map = {}
            for col_no, val in enumerate(row_values):
                if val in EXPECTED_HEADERS:
                    input_col_map[val] = col_no

            if "Supplier" in input_col_map and "DocumentNo" in input_col_map:
                positions.append((idx, input_col_map))

    return positions


def is_subtotal_or_total_row(raw_row):
    first_cell = normalize_text(raw_row.iloc[0] if len(raw_row) > 0 else "")
    return first_cell in ("*", "**")


def read_new_sapos_excel(input_path, sheet_name=None):
    """Read the new repeated-header Excel file and return clean transaction rows."""
    if not Path(input_path).exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    raw_df = pd.read_excel(input_path, sheet_name=sheet_name or 0, header=None, dtype=object)
    header_positions = find_header_positions(raw_df)

    if not header_positions:
        raise Exception("No SAP header row found. Expected repeated header with Supplier, Vendor Name, DocumentNo.")

    output_rows = []
    block_no = 0

    for pos_no, (header_row_idx, input_col_map) in enumerate(header_positions):
        block_no += 1
        next_header_idx = header_positions[pos_no + 1][0] if pos_no + 1 < len(header_positions) else len(raw_df)

        start_row = header_row_idx + 1
        end_row = next_header_idx

        for ridx in range(start_row, end_row):
            raw_row = raw_df.iloc[ridx]

            if is_subtotal_or_total_row(raw_row):
                continue

            supplier = normalize_supplier_code(raw_row.iloc[input_col_map["Supplier"]])
            document_no = normalize_reference(raw_row.iloc[input_col_map["DocumentNo"]])

            # Real transaction row must have supplier and document number.
            if not supplier or not document_no:
                continue
            if not is_valid_supplier_code(supplier):
                continue

            vendor_name = normalize_text(raw_row.iloc[input_col_map.get("Vendor Name", -1)]).upper()
            pmnt_dt = parse_date_safe(raw_row.iloc[input_col_map.get("Pmnt Date", -1)])
            text_value = normalize_text(raw_row.iloc[input_col_map.get("Text", -1)])
            dd_value = normalize_text(raw_row.iloc[input_col_map.get("DD", -1)])
            real_due = parse_real_due_dt(text_value, pmnt_dt, dd_value)

            output_rows.append({
                "BlockNo": block_no,
                "Supplier": supplier,
                "Name": vendor_name,
                "City": "",  # New Excel format does not carry City separately.
                "DocumentNo": document_no,
                "St": normalize_text(raw_row.iloc[input_col_map.get("St", -1)]),
                "Reference": normalize_reference(raw_row.iloc[input_col_map.get("Reference", -1)]),
                "PK": normalize_text(raw_row.iloc[input_col_map.get("PK", -1)]),
                "Type": normalize_text(raw_row.iloc[input_col_map.get("Type", -1)]),
                "DD": dd_value,
                "Doc. Date": parse_date_safe(raw_row.iloc[input_col_map.get("Doc. Date", -1)]),
                "Pstng Date": parse_date_safe(raw_row.iloc[input_col_map.get("Pstng Date", -1)]),
                "S": normalize_text(raw_row.iloc[input_col_map.get("S", -1)]),
                "Pmnt date": pmnt_dt,
                "RealDue_Dt": real_due,
                "Cash Flow": cashflow_bucket(real_due),
                "LC amnt": clean_amount(raw_row.iloc[input_col_map.get("Local Crcy Amt", -1)]),
                "LCu": normalize_text(raw_row.iloc[input_col_map.get("LCu", -1)]),
                "Amount in DC": clean_amount(raw_row.iloc[input_col_map.get("Amount in DC", -1)]),
                "Crcy": normalize_text(raw_row.iloc[input_col_map.get("Crcy", -1)]),
                "Text": text_value,
                "User Name": normalize_text(raw_row.iloc[input_col_map.get("User Name", -1)]),
                "Clearing": normalize_text(raw_row.iloc[input_col_map.get("Clearing", -1)]),
            })

    return pd.DataFrame(output_rows)


# ================== OUTPUT FORMATTING ==================

def autosize_columns(writer, sheet_name, df):
    """Basic Excel column width adjustment."""
    ws = writer.book[sheet_name]
    for idx, col in enumerate(df.columns, 1):
        max_len = len(str(col))
        for val in df[col].head(5000):
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = min(max_len + 2, 45)


def write_final_output(df_out, output_path):
    output_folder = Path(output_path).parent
    output_folder.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl", date_format="DD-MM-YYYY", datetime_format="DD-MM-YYYY") as writer:
        df_out.to_excel(writer, index=False, sheet_name="SAP_OS")
        autosize_columns(writer, "SAP_OS", df_out)


# ================== MAIN ==================

def main():
    print("Reading master file...")
    wb_m, ws_m, col_map, group_map, view_map, name_map, existing = read_master_maps(MASTER_PATH)

    print("Reading new SAP OS Excel format...")
    df_out = read_new_sapos_excel(INPUT_PATH, INPUT_SHEET_NAME)

    if df_out.empty:
        raise Exception("No transaction rows found in input file.")

    print("Applying master data Group / View - Y/N...")
    df_out.insert(4, "Group", df_out["Supplier"].map(group_map).fillna(""))
    df_out.insert(5, "View - Y/N", df_out["Supplier"].map(view_map).fillna(""))

    print("Checking missing suppliers in master...")
    missing = {}
    for _, row in df_out[["Supplier", "Name"]].drop_duplicates().iterrows():
        sup = normalize_supplier_code(row["Supplier"])
        name = normalize_text(row["Name"])
        if sup and sup not in existing:
            missing[sup] = name

    if missing:
        today = datetime.today().date()
        append_missing_to_master(ws_m, col_map, [(sup, name, today) for sup, name in missing.items()])
        wb_m.save(MASTER_PATH)
        print(f"Added {len(missing)} missing supplier(s) to master file.")
    else:
        print("No missing supplier found.")

    print("Writing final output Excel...")
    write_final_output(df_out, OUTPUT_PATH)

    print("SUCCESS")
    print(f"Rows processed: {len(df_out)}")
    print(f"Output file: {OUTPUT_PATH}")

    try:
        os.startfile(OUTPUT_PATH)  # Windows only
    except Exception:
        pass


if __name__ == "__main__":
    main()
